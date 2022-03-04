#########################
##### Version 0.1.3 #####
#########################

import os
import tkinter as tk
from tkinter import ttk
from tkinter import filedialog
from tkinter import messagebox
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
import datetime
import jpholiday

dir_name = os.path.abspath(os.path.dirname(__file__))
week_list = ["月", "火", "水", "木", "金", "土", "日"]


def button1_clicked():
    file_path = filedialog.askopenfilename(
        initialdir=dir_name, filetypes=[("Excel Files", "*xlsx")]
    )
    if file_path:
        name1.set(file_path)


def button2_clicked():
    file_path = filedialog.askopenfilename(
        initialdir=dir_name, filetypes=[("Excel Files", "*xlsx")]
    )
    if file_path:
        name2.set(file_path)


def button3_clicked():
    schedule_book_name = name1.get()
    # schedule_book_name = "./xlsx_files/schedule/01_schedule.xlsx"
    # schedule_book_name = "./xlsx_files/schedule/02_schedule_excess_sheet.xlsx"
    # schedule_book_name = "./xlsx_files/schedule/03_schedule_values_at_space.xlsx"
    # schedule_book_name = "./xlsx_files/schedule/04_schedule_duplicate.xlsx"
    # schedule_book_name = "./xlsx_files/schedule/05_schedule_two_rows_in_a_day.xlsx"
    # schedule_book_name = "./xlsx_files/schedule/06_schedule_two_rows_in_two_days.xlsx"
    # schedule_book_name = "./xlsx_files/schedule/07_schedule_short.xlsx"
    # schedule_book_name = "./xlsx_files/schedule/08_schedule_blank.xlsx"
    # schedule_book_name = "./xlsx_files/schedule/09_schedule_diff_in_names.xlsx"

    template_book_name = name2.get()
    # template_book_name = "./xlsx_files/template/template.xlsx"

    schedule_book = openpyxl.load_workbook(schedule_book_name)
    template_book = openpyxl.load_workbook(template_book_name)

    schedule_sheet_names = schedule_book.sheetnames
    template_sheet_names = template_book.sheetnames
    if len(schedule_sheet_names) > 1 or len(template_sheet_names) > 1:
        messagebox.showerror("エラー", "シートは1つにしてください。")
    else:
        try:
            df = pd.read_excel(schedule_book_name, sheet_name=0)

            df_with_new_rows = df
            for index, row in df.iterrows():
                new_work_title = ""
                if row[1] == "当日直":
                    df_with_new_rows.iloc[index, 1] = "当直"
                    new_work_title = "日直"
                elif row[1] == "日当日直":
                    df_with_new_rows.iloc[index, 1] = "日当直"
                    new_work_title = "日直"
                elif row[1] == "当直＋午前":
                    df_with_new_rows.iloc[index, 1] = "当直"
                    new_work_title = "午前"
                if new_work_title != "":
                    new_row = pd.Series(
                        [
                            row[0],
                            new_work_title,
                            row[2],
                            row[6],
                            row[4],
                            row[5],
                            row[6],
                            row[7],
                            row[8],
                        ],
                        index=df.columns,
                    )
                    df_with_new_rows = df_with_new_rows.append(
                        new_row, ignore_index=True
                    )

            df_with_new_rows["year"] = (
                df_with_new_rows["開始日"].dt.strftime("%Y").map(lambda str: int(str))
            )
            df_with_new_rows["month"] = (
                df_with_new_rows["開始日"].dt.strftime("%m").map(lambda str: int(str))
            )
            df_with_new_rows["date"] = df_with_new_rows["開始日"].dt.day
            df_with_new_rows["week_index"] = df_with_new_rows["開始日"].dt.weekday
            df_with_new_rows["weekday"] = df_with_new_rows["week_index"].map(
                lambda index: week_list[index]
            )
            df_with_new_rows["doctor"] = df_with_new_rows["担当"]
            df_with_new_rows["value"] = df_with_new_rows[["施設名", "仕事内容"]].apply(
                " ".join, axis=1
            )
            df_wide = df_with_new_rows.pivot(
                index=["year", "month", "date", "weekday"],
                columns="doctor",
                values="value",
            )
            df_wide = df_wide.rename_axis().reset_index()

            date_count = df_wide.shape[0]
            year = df_wide["year"].mode().tolist()[0]
            month = df_wide["month"].mode().tolist()[0]
            doctors_in_schedule = df["担当"].unique().tolist()

            ws = template_book[template_sheet_names[0]]
            ws["B2"].value = "ネーベン表 " + str(year) + " 年 " + str(month) + " 月"

            # fit size according to date_count
            num_row = ws.max_row
            num_del_rows = num_row - 6 - date_count
            if num_del_rows > 0:
                ws.delete_rows(idx=6, amount=num_del_rows)

            # write date and weekday values according to the schedule sheet information
            for i in range(date_count):
                # write date (and month) values
                if df_wide["month"][i] != month or df_wide["date"][i] == 1:
                    ws.cell(5 + i, 2).value = (
                        str(df_wide["month"][i]) + "/" + str(df_wide["date"][i])
                    )
                else:
                    ws.cell(5 + i, 2).value = str(df_wide["date"][i])
                ws.cell(5 + i, 3).value = df_wide["weekday"][i]

            # write doctor names and change background filling at null cells
            num_col = ws.max_column
            for i in range(4, num_col - 2):
                doctor = ws.cell(4, i).value
                if doctor in doctors_in_schedule:
                    for j in range(date_count):
                        value = df_wide[doctor][j]
                        ws.cell(5 + j, i).value = value
                        if pd.isna(value):
                            ws.cell(5 + j, i).fill = PatternFill(fill_type=None)

            # check if doctor names in schedule sheet are indeed in doctors_in_template
            doctors_in_template = []
            for i in range(4, num_col - 2):
                doctors_in_template.append(ws.cell(4, i).value)
            doctors_missing = []
            for name in doctors_in_schedule:
                if name not in doctors_in_template:
                    doctors_missing.append(str(name))

            # change color the row where Saturday, Sunday or holiday
            for i in range(date_count):
                weekday = df_wide["weekday"][i]
                date = datetime.datetime(
                    df_wide["year"][i], df_wide["month"][i], df_wide["date"][i]
                )
                if weekday == "日" or jpholiday.is_holiday(date):
                    for j in range(2, num_col):
                        ws.cell(5 + i, j).fill = PatternFill(
                            patternType="solid", fgColor="ff3333"
                        )
                if weekday == "土":
                    for j in range(2, num_col):
                        ws.cell(5 + i, j).fill = PatternFill(
                            patternType="solid", fgColor="ffff00"
                        )

            # generate timestamp for created xlsx file name
            dt_now = datetime.datetime.now(
                datetime.timezone(datetime.timedelta(hours=9))
            )
            timestamp = dt_now.strftime("%Y%m%d%H%M%S")
            dest_path = os.path.join(
                dir_name, "schedule2tablexloutput_" + timestamp + ".xlsx"
            )
            template_book.save(dest_path)

            message = dest_path + "を作成しました\n"
            is_nan_bool_list = list(map(lambda str: str == "nan", doctors_missing))
            if sum(is_nan_bool_list) > 0:
                message = message + "空欄があります\n"
            doctors_missing_without_nan = [x for x in doctors_missing if x != "nan"]
            if len(doctors_missing_without_nan) > 0:
                message = (
                    message + ",".join(doctors_missing_without_nan) + "先生がうまく登録できていません"
                )

            messagebox.showinfo("お知らせ", message)

        except ValueError:
            messagebox.showerror("エラー", "うまく処理できません")


if __name__ == "__main__":
    root = tk.Tk()
    root.title("当直表作りお助けプログラム")
    root.geometry("600x150")

    frame1 = ttk.Frame(root, padding=10)
    frame1.grid()

    set1 = tk.StringVar()
    set1.set("スケジュールファイル: ")
    label1 = ttk.Label(frame1, textvariable=set1)
    label1.grid(row=0, column=0)

    name1 = tk.StringVar()
    entry1 = ttk.Entry(frame1, textvariable=name1, width=50)
    entry1.grid(row=0, column=1)

    button1 = ttk.Button(frame1, text="参照", command=button1_clicked)
    button1.grid(row=0, column=2)

    frame2 = ttk.Frame(root, padding=10)
    frame2.grid()

    set2 = tk.StringVar()
    set2.set("テンプレートファイル: ")
    label2 = ttk.Label(frame2, textvariable=set2)
    label2.grid(row=0, column=0)

    name2 = tk.StringVar()
    entry2 = ttk.Entry(frame2, textvariable=name2, width=50)
    entry2.grid(row=0, column=1)

    button2 = ttk.Button(frame2, text="参照", command=button2_clicked)
    button2.grid(row=0, column=2)

    frame3 = ttk.Frame(root, padding=10)
    frame3.grid()

    button3 = ttk.Button(frame3, text="実行", command=button3_clicked)
    button3.grid(row=0, column=0)

    root.mainloop()

# button3_clicked()
